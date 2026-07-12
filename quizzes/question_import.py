"""Bulk question import from CSV/XLSX (Part G).

No pandas — csv (stdlib) for .csv, openpyxl for .xlsx. Parsing and
row-validation are kept separate from the view so they can be unit
tested / reused independently of the request/response cycle.
"""
import csv
import io

from openpyxl import load_workbook

REQUIRED_COLUMNS = ["question", "option_1", "option_2", "option_3", "option_4", "correct_option"]
OPTIONAL_COLUMNS = ["explanation"]

# 1/2/3/4, A/B/C/D, क/ख/ग/घ all normalise to the model's 1-4 IntegerField.
_OPTION_MAP = {
    "1": 1, "2": 2, "3": 3, "4": 4,
    "a": 1, "b": 2, "c": 3, "d": 4,
    "क": 1, "ख": 2, "ग": 3, "घ": 4,
}


def normalize_correct_option(raw):
    key = str(raw if raw is not None else "").strip().lower()
    return _OPTION_MAP.get(key)


def parse_upload(uploaded_file):
    """Returns a list of dict rows (one per data row) regardless of source format."""
    name = (uploaded_file.name or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(uploaded_file)
    return _parse_csv(uploaded_file)


def _parse_csv(uploaded_file):
    text = uploaded_file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(uploaded_file):
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
    out = []
    for r in rows_iter:
        if all(c is None for c in r):
            continue
        out.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
    return out


def _clean_str(v):
    if v is None:
        return ""
    return v.strip() if isinstance(v, str) else str(v).strip()


def validate_rows(raw_rows, existing_titles_lower):
    """Returns (clean_rows, errors, warnings).

    clean_rows: list of dicts ready for Questions(quiz=quiz, **row).
    errors: list of (row_number, message) -- ANY error rejects the WHOLE
            import (transaction.atomic() in the view enforces all-or-nothing).
    warnings: list of (row_number, message) -- duplicates, non-fatal.
    row_number counts from 2 (row 1 is the header), matching what a human
    looking at the spreadsheet would call that row.
    """
    errors = []
    warnings = []
    clean_rows = []
    seen_in_file = set()

    if not raw_rows:
        errors.append((0, "No data rows found in the file."))
        return [], errors, warnings

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in raw_rows[0]]
    if missing_cols:
        errors.append((0, f"Missing column(s): {', '.join(missing_cols)} — download the template to check the format."))
        return [], errors, warnings

    for idx, row in enumerate(raw_rows, start=2):
        q_text = _clean_str(row.get("question"))
        opts = [_clean_str(row.get(f"option_{n}")) for n in (1, 2, 3, 4)]
        correct_raw = row.get("correct_option")
        explanation = _clean_str(row.get("explanation"))

        if not q_text:
            errors.append((idx, "Question text is empty."))
            continue
        if any(not o for o in opts):
            errors.append((idx, "One or more options are empty (option_1 through option_4 are all required)."))
            continue

        correct_int = normalize_correct_option(correct_raw)
        if correct_int is None:
            errors.append((idx, f"correct_option '{correct_raw}' not recognized — use one of 1/2/3/4, A/B/C/D, or क/ख/ग/घ."))
            continue

        key = q_text.lower()
        if key in seen_in_file or key in existing_titles_lower:
            warnings.append((idx, f"Duplicate question (already in this quiz): \"{q_text[:60]}\""))
        seen_in_file.add(key)

        clean_rows.append({
            "question": q_text,
            "option_1": opts[0], "option_2": opts[1], "option_3": opts[2], "option_4": opts[3],
            "correct_option": correct_int,
            "explanation": explanation,
        })

    return clean_rows, errors, warnings


SAMPLE_ROWS = [
    ["question", "option_1", "option_2", "option_3", "option_4", "correct_option", "explanation"],
    ["What is the capital of India?", "Mumbai", "Delhi", "Kolkata", "Chennai", "2", "Delhi is the capital of India."],
    ["What is 2 + 2?", "3", "4", "5", "6", "B", ""],
]
